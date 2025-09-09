from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

app = Flask(__name__)
CORS(app)  # Enable CORS so React frontend can access Flask API

COLUMNS = ["ID", "Name", "Email", "Phone", "Address", "City", "State", "Zip", "Country", "Status"]

UPLOAD_FOLDER = "generated_files"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def normalize_nulls(df):
    """Replace empty strings and NaNs with None."""
    return df.where(pd.notnull(df), None).replace(r'^\s*$', None, regex=True)

def highlight_differences(df_benchmark, df_new):
    """Find unique rows and highlight changed columns."""
    df_benchmark = normalize_nulls(df_benchmark)
    df_new = normalize_nulls(df_new)

    merged = pd.merge(df_benchmark, df_new, how="outer", indicator=True)
    unique_rows = merged[merged["_merge"] != "both"].drop(columns=["_merge"])

    change_summary = []
    for _, row in unique_rows.iterrows():
        matching_benchmark = df_benchmark[df_benchmark["ID"] == row["ID"]]
        if not matching_benchmark.empty:
            benchmark_row = matching_benchmark.iloc[0]
            changes = []
            for col in COLUMNS:
                val_new = row[col]
                val_bench = benchmark_row[col]
                if (val_new is None and val_bench is None):
                    continue
                elif val_new != val_bench:
                    changes.append(col)
            change_summary.append(", ".join(changes) if changes else "New Entry")
        else:
            change_summary.append("New Entry")

    unique_rows["Change Summary"] = change_summary
    return unique_rows

def save_with_highlights(df, filename):
    """Save unique rows to Excel with highlighted differences."""
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_idx in range(2, ws.max_row + 1):
        change_cols = ws.cell(row=row_idx, column=len(COLUMNS) + 1).value
        if change_cols and change_cols != "New Entry":
            cols_to_highlight = change_cols.split(", ")
            for col_name in cols_to_highlight:
                if col_name in COLUMNS:
                    col_idx = COLUMNS.index(col_name) + 1
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    file_path = os.path.join(UPLOAD_FOLDER, filename)
    wb.save(file_path)
    return file_path

@app.route("/compare", methods=["POST"])
def compare_files():
    benchmark_file = request.files.get("benchmark")
    new_data_file = request.files.get("new_data")

    if not benchmark_file or not new_data_file:
        return jsonify({"error": "Please upload both files"}), 400

    df_benchmark = pd.read_excel(benchmark_file, usecols=COLUMNS)
    df_new = pd.read_excel(new_data_file, usecols=COLUMNS)

    df_benchmark = normalize_nulls(df_benchmark)
    df_new = normalize_nulls(df_new)

    duplicates = pd.merge(df_benchmark, df_new, how="inner")
    duplicates_path = os.path.join(UPLOAD_FOLDER, "duplicates.xlsx")
    duplicates.to_excel(duplicates_path, index=False)

    unique_entities = highlight_differences(df_benchmark, df_new)
    unique_path = save_with_highlights(unique_entities, "unique_highlighted.xlsx")

    return jsonify({
        "duplicates_url": f"/download/duplicates.xlsx",
        "unique_url": f"/download/unique_highlighted.xlsx"
    })

@app.route("/download/<filename>")
def download_file(filename):
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return jsonify({"error": "File not found"}), 404

if __name__ == "__main__":
    app.run(debug=True)
