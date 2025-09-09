from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

app = Flask(__name__)
CORS(app)  # Enable CORS so React frontend can access Flask API

COLUMNS = ["ID", "Name", "Age", "City"]
UPLOAD_FOLDER = "generated_files"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def normalize_nulls(df):
    """Replace empty strings and NaNs with None."""
    return df.where(pd.notnull(df), None).replace(r'^\s*$', None, regex=True)

def find_duplicates_and_unique(df_benchmark, df_new):
    """Find duplicates and unique rows between benchmark and new data."""
    df_benchmark = normalize_nulls(df_benchmark)
    df_new = normalize_nulls(df_new)
    
    # Find exact duplicates (rows that exist in both files with same values)
    duplicates = pd.merge(df_benchmark, df_new, how="inner", on=COLUMNS)
    
    # Find all unique rows (including modified and completely new)
    merged = pd.merge(df_benchmark, df_new, how="outer", indicator=True, on=COLUMNS)
    
    # Rows only in benchmark (removed in new data)
    only_in_benchmark = merged[merged["_merge"] == "left_only"].drop(columns=["_merge"])
    
    # Rows only in new data (completely new entries)
    only_in_new = merged[merged["_merge"] == "right_only"].drop(columns=["_merge"])
    
    # Find modified rows (same ID but different other values)
    modified_rows = []
    
    # Get IDs that exist in both files
    benchmark_ids = set(df_benchmark['ID'].dropna())
    new_ids = set(df_new['ID'].dropna())
    common_ids = benchmark_ids.intersection(new_ids)
    
    for id_val in common_ids:
        bench_rows = df_benchmark[df_benchmark['ID'] == id_val]
        new_rows = df_new[df_new['ID'] == id_val]
        
        if len(bench_rows) == 1 and len(new_rows) == 1:
            bench_row = bench_rows.iloc[0]
            new_row = new_rows.iloc[0]
            
            # Check if any non-ID columns are different
            is_different = False
            changes = []
            for col in COLUMNS:
                if col != 'ID':
                    if bench_row[col] != new_row[col]:
                        is_different = True
                        changes.append(col)
            
            if is_different:
                new_row_dict = new_row.to_dict()
                new_row_dict['Change Summary'] = ", ".join(changes)
                modified_rows.append(new_row_dict)
    
    # Combine all unique rows
    unique_rows = []
    
    # Add modified rows
    unique_rows.extend(modified_rows)
    
    # Add completely new rows
    for _, row in only_in_new.iterrows():
        row_dict = row.to_dict()
        row_dict['Change Summary'] = "New Entry"
        unique_rows.append(row_dict)
    
    # Add removed rows (from benchmark, not in new data)
    for _, row in only_in_benchmark.iterrows():
        row_dict = row.to_dict()
        row_dict['Change Summary'] = "Removed Entry"
        unique_rows.append(row_dict)
    
    unique_df = pd.DataFrame(unique_rows) if unique_rows else pd.DataFrame(columns=COLUMNS + ['Change Summary'])
    
    return duplicates, unique_df

def save_with_highlights(df, filename):
    """Save unique rows to Excel with highlighted differences."""
    if df.empty:
        # Create empty file with headers
        buffer = io.BytesIO()
        pd.DataFrame(columns=COLUMNS + ['Change Summary']).to_excel(buffer, index=False)
        buffer.seek(0)
        
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(file_path, 'wb') as f:
            f.write(buffer.getvalue())
        return file_path
    
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    
    wb = load_workbook(buffer)
    ws = wb.active
    
    # Yellow fill for highlighting
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Highlight changes
    for row_idx in range(2, ws.max_row + 1):
        change_cols = ws.cell(row=row_idx, column=len(COLUMNS) + 1).value
        if change_cols:
            if change_cols == "New Entry" or change_cols == "Removed Entry":
                # Highlight entire row for new/removed entries
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
            else:
                # Highlight specific changed columns
                cols_to_highlight = change_cols.split(", ")
                for col_name in cols_to_highlight:
                    if col_name in COLUMNS:
                        col_idx = COLUMNS.index(col_name) + 1
                        ws.cell(row=row_idx, column=col_idx).fill = fill
    
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    wb.save(file_path)
    return file_path

@app.route("/compare", methods=["POST"])
def compare_files():
    benchmark_file = request.files.get("benchmark")
    new_data_file = request.files.get("new_data")
    
    if not benchmark_file or not new_data_file:
        return jsonify({"error": "Please upload both files"}), 400
    
    try:
        # Read Excel files
        df_benchmark = pd.read_excel(benchmark_file, usecols=COLUMNS)
        df_new = pd.read_excel(new_data_file, usecols=COLUMNS)
        
        # Normalize data
        df_benchmark = normalize_nulls(df_benchmark)
        df_new = normalize_nulls(df_new)
        
        # Find duplicates and unique entities
        duplicates, unique_entities = find_duplicates_and_unique(df_benchmark, df_new)
        
        # Save duplicates file
        duplicates_path = os.path.join(UPLOAD_FOLDER, "duplicates.xlsx")
        duplicates.to_excel(duplicates_path, index=False)
        
        # Save unique entities with highlights
        unique_path = save_with_highlights(unique_entities, "unique_highlighted.xlsx")
        
        # Calculate statistics
        stats = {
            "benchmark_rows": len(df_benchmark),
            "new_rows": len(df_new),
            "duplicates_count": len(duplicates),
            "unique_count": len(unique_entities)
        }
        
        return jsonify({
            "duplicates_url": f"/download/duplicates.xlsx",
            "unique_url": f"/download/unique_highlighted.xlsx",
            "stats": stats
        })
        
    except Exception as e:
        return jsonify({"error": f"Error processing files: {str(e)}"}), 500

@app.route("/download/<filename>")
def download_file(filename):
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return jsonify({"error": "File not found"}), 404

if __name__ == "__main__":
    app.run(debug=True)